from collections import defaultdict
from io import BytesIO
from openpyxl.styles import Alignment, Border, Font, Side
import hashlib
import logging
import time

import pandas as pd
from django.conf import settings
from django.core.cache import cache
from django.http import FileResponse
from uuid import uuid4
from zipfile import ZipFile, ZIP_DEFLATED


from common.utils.yadisk import YandexDiskUploader
from questionnaire.models import Survey


logger = logging.getLogger(__name__)

DOWNLOAD_URL_ERROR = "Ошибка при получении URL для скачивания файла: {}"
DOWNLOAD_ERROR = "Ошибка загрузки документа {}: {}"


def get_url(document):
    """Получение URL на скачивание файла от API Yandex-диска."""
    try:
        uploader = YandexDiskUploader(settings.DISK_TOKEN)
        download_url = uploader.get_download_url(document.image)
        if download_url and download_url != "#":
            return download_url
    except Exception as e:
        logger.error(DOWNLOAD_URL_ERROR.format(e))


def get_cached_yadisk_url(file_path, ttl_seconds=1500):
    """Получение URL Яндекс-диска с кешированием."""
    if not file_path:
        return None

    cache_key = f"yadisk_url_{hashlib.md5(file_path.encode()).hexdigest()}"

    # Пробуем получить из кеша
    cached = cache.get(cache_key)
    if cached and time.time() < cached['expires']:
        return cached['url']

    # Получаем новую ссылку
    class TempDoc:
        def __init__(self, file_path):
            self.image = file_path

    temp_doc = TempDoc(file_path)
    download_url = get_url(temp_doc)

    if download_url and download_url != "#":
        # Кешируем на указанное время
        cache.set(cache_key, {
            'url': download_url,
            'expires': time.time() + ttl_seconds
        }, ttl_seconds)

    return download_url if download_url and download_url != "#" else None


def get_excel_file(queryset):
    """Формирование Excel-файла."""

    # Группировка данных с использованием defaultdict
    group_by_uuid = defaultdict(list)

    for survey in queryset.select_related('user'):
        # Формируем базовую информацию
        survey_result = {
            "ФИО пользователя": f"{survey.user.first_name} "
                                f"{survey.user.last_name} "
                                f"{survey.user.patronymic}".strip()
        }

        # Преобразуем результаты в словарь более эффективно
        result_dict = dict(zip(survey.result[::2], survey.result[1::2]))
        survey_result.update(result_dict)

        group_by_uuid[survey.questions_version_uuid].append(survey_result)

    # Создаем Excel файл в памяти
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for uuid, results in group_by_uuid.items():
            # Создаем DataFrame из результатов и транспонируем
            df = pd.DataFrame(results)
            df_transposed = df.T  # Транспонирование

            # Сохраняем в Excel без индексов и заголовков
            sheet_name = str(uuid)[:31]
            df_transposed.to_excel(
                writer,
                sheet_name=sheet_name,
                index=True,  # Сохраняем индексы (вопросы)
                header=False,  # Не используем автоматические заголовки
                engine='openpyxl'
            )

            # Получаем рабочий лист для настройки
            worksheet = writer.sheets[sheet_name]

            # Стиль для рамки
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Стиль для выравнивания и переноса текста
            wrap_align = Alignment(
                wrap_text=True,
                vertical='top',
                horizontal='left'  # Выравнивание по левому краю
            )

            # Стиль для заголовков
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(
                wrap_text=True,
                vertical='center',
                horizontal='center'
            )

            # Устанавливаем заголовки столбцов
            worksheet.cell(row=1, column=1, value="Вопрос")
            worksheet.cell(row=1, column=2, value="Ответ")

            # Применяем стили к заголовкам
            for col in [1, 2]:
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Настраиваем ширину столбцов
            # Столбец "Вопрос" - 100
            worksheet.column_dimensions['A'].width = 100
            # Столбец "Ответ" - 80
            worksheet.column_dimensions['B'].width = 80

            # Применяем стили ко всем ячейкам
            # Начинаем с 2 строки, т.к. 1 - заголовки
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = wrap_align
                    cell.border = thin_border

            # Применяем рамку к заголовкам
            for col in [1, 2]:
                cell = worksheet.cell(row=1, column=col)
                cell.border = thin_border

    # Подготавливаем ответ
    output.seek(0)
    file_name = f"survey_report_{uuid4()}.xlsx"

    response = FileResponse(
        output,
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    return response


def get_docs_zip(request, uuid):
    """Упаковка документов в zip-файл в памяти."""
    import requests
    from io import BytesIO

    survey = Survey.objects.get(id=uuid)
    documents = survey.docs.all()

    # Создаем zip-файл в памяти
    zip_buffer = BytesIO()

    with ZipFile(
            zip_buffer,
            "w",
            compression=ZIP_DEFLATED, compresslevel=6
    ) as zip_file:
        for document in documents:
            try:
                response = requests.get(get_url(document), timeout=30)
                if response.status_code == 200:
                    extension = (
                        document.image.split('.')[-1]
                        if '.' in document.image else 'jpg'
                    )
                    image_name = (
                        f"{document.id}_{uuid4().hex[:8]}.{extension}"
                    )

                    # Добавляем файл в zip напрямую из памяти
                    zip_file.writestr(image_name, response.content)
            except requests.RequestException as e:
                print(DOWNLOAD_ERROR.format(document.id, e))
                continue

    # Подготавливаем ответ
    zip_buffer.seek(0)
    file_name = f"survey_{uuid}_documents.zip"

    response = FileResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    response["Content-Length"] = zip_buffer.getbuffer().nbytes

    return response

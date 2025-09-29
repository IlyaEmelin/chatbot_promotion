import pathlib
import tempfile
import os
from urllib.parse import urlparse

from django.contrib import admin
from django.contrib.auth import get_user_model
from django.http import FileResponse
from django.utils import timezone
from django.utils.html import format_html
from openpyxl import Workbook

from questionnaire.constant import STATUS_CHOICES
from questionnaire.models import AnswerChoice, Document, Survey, Question

User = get_user_model()


class StatusFilter(admin.SimpleListFilter):
    """Фильтр по статусу опроса."""

    title = "Статус заявки"
    parameter_name = "status"

    def lookups(self, request, model_admin):
        return STATUS_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(status=self.value())
        return queryset


class DocumentInline(admin.TabularInline):
    """Документы."""

    model = Document
    extra = 0
    readonly_fields = ("image_preview", "download_link")

    @admin.display(description="Предпросмотр")
    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<a href="{}" target="_blank"><img src="{}"'
                'style="max-height: 100px; max-width: 100px;" /></a>',
                obj.image,
                obj.image,
            )
        return "—"

    @admin.display(description="Скачать")
    def download_link(self, obj):
        if obj.image:
            return format_html(
                '<a href="{}" target="_blank" download>' "Скачать</a>",
                obj.image,
            )
        return "—"


@admin.register(Survey)
class SurveyAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "user_info",
        "status",
        "documents_count",
        "created_at_formatted",
    )
    list_filter = (StatusFilter, "created_at")
    search_fields = ("user__email", "user__phone_number", "created_at")
    readonly_fields = (
        "id",
        "user_info",
        "result_display",
        "current_question",
        "created_at",
        "updated_at",
        "questions_version_uuid",
        "documents_list",
    )
    exclude = ("user", "result",)
    list_select_related = ("user", "current_question")
    inlines = (DocumentInline,)
    actions = ['download_servey']

    @admin.action(description='Скачать результаты опроса в формате Excel')
    def download_servey(self, request, queryset):

        group_by_uuid = {}
        for servey in queryset:
            servey_result = {
                'ФИО пользователя': f'{servey.user.first_name} '
                f'{servey.user.last_name} '
                f'{servey.user.patronymic}'
            }
            # переводим резульаты из списка в словарь
            servey_result.update({
                servey.result[i]: servey.result[i+1] for i in range(
                    0, len(servey.result)-1, 2)
            })
            group_by_uuid.setdefault(
                servey.questions_version_uuid, []
            ).append(servey_result)

        with tempfile.TemporaryDirectory() as report_dir:
            path_dir = pathlib.Path(report_dir)

            servey_report = Workbook()
            for uuid, results in group_by_uuid.items():
                sheet = servey_report.create_sheet('Mysheet', 0)
                sheet.title = f'{uuid}'
                row = 1
                # заполняем заголовки из первого опроса
                for question in results[0].keys():
                    sheet.cell(row=row, column=1, value=question)
                    row += 1
                column = 2
                # для каждого опроса
                for result in results:
                    row = 1
                    # заполняем результаты
                    for answer in result.values():
                        sheet.cell(row=row, column=column, value=answer)
                        row += 1
                    column += 1

            servey_report.save(path_dir/'servey_report.xlsx')

            response = FileResponse(
                open(path_dir/'servey_report.xlsx', 'rb')
            )
            return response

    @admin.display(description="Пользователь")
    def user_info(self, obj):
        user = obj.user
        phone = getattr(user, "phone", "—")
        return format_html(
            "{} {}<br><small>Почта: {}<br>" "Телефон: {}</small>",
            user.first_name or "",
            user.last_name or "",
            user.email,
            phone,
        )

    # @admin.display(description="Статус")
    # def status_badge(self, obj):
    #     status_colors = {
    #         "draft": "gray",
    #         "new": "blue",
    #         "waiting_docs": "orange",
    #         "processing": "purple",
    #         "completed": "green",
    #     }
    #     color = status_colors.get(obj.status, "gray")
    #     return format_html(
    #         '<span style="background-color: {}; color: '
    #         "white; padding: 4px 8px; border-radius: "
    #         '12px; font-size: 12px;">{}</span>',
    #         color,
    #         obj.get_status_display(),
    #     )

    @admin.display(description="Создана")
    def created_at_formatted(self, obj):
        return timezone.localtime(obj.created_at).strftime("%d.%m.%Y %H:%M")

    @admin.display(description="Документы")
    def documents_count(self, obj):
        return obj.docs.count()

    @admin.display(description="Результаты опроса")
    def result_display(self, obj):
        result = ''
        for i in range(0, len(obj.result), 2):
            result += f'Вопрос: {obj.result[i]}\nОтвет: {obj.result[i+1]}\n'
        return result

    @admin.display(description="Загруженные документы")
    def documents_list(self, obj):
        documents = obj.docs.all()
        if not documents:
            return "Документы не загружены"

        html = '<div style="max-height: 200px; overflow-y: auto;">'
        for doc in documents:
            filename = os.path.basename(urlparse(doc.image).path)
            html += format_html(
                """
                <div style="display: flex; align-items: '
                'center; margin-bottom: 10px; padding: 5px; background: '
                'white; border-radius: 3px;">
                    <a href="{}" target="_blank" style="margin-right: 10px;">
                        <img src="{}" style="max-height: 50px; '
                        'max-width: 50px;">
                    </a>
                    <div>
                        <div><strong>{}</strong></div>
                        <a href="{}" target="_blank" download>Скачать</a>
                    </div>
                </div>
                """,
                doc.image,
                doc.image,
                filename,
                doc.image,
            )
        html += "</div>"
        return format_html(html)


@admin.register(Document)
class DocumentAdmin(admin.ModelAdmin):
    """Документ."""

    list_display = ("survey_short", "image_preview")

    @admin.display(description="Опрос")
    def survey_short(self, obj):
        return f"Опрос {obj.survey.id}"

    @admin.display(description="Изображение")
    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<a href="{}" target="_blank"><img src="{}" '
                'style="max-height: 50px;" /></a>',
                obj.image,
                obj.image,
            )
        return "—"


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    """Вопрос."""

    list_display = ("text", "type")


@admin.register(AnswerChoice)
class AnswerChoiceAdmin(admin.ModelAdmin):
    """Вопрос."""

    list_display = ("current_question", "next_question", "answer")


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    model = User
    fieldsets = (
        (None, {
            'fields': (
                'username',
                'password'
            )
        }),
        ('Персональная информация', {
            'fields': (
                'first_name',
                'last_name',
                'patronymic',
                'email',
                'phone_number',
                'telegram_username',
                'residence'
            )
        }),
        ('Подопечный', {
            'fields': (
                'agent_status',
                'ward_first_name',
                'ward_last_name',
                'ward_patronymic',
                'birthday',
            )
        }),
        ('Статус', {
            'fields': (
                'is_active',
                'is_superuser',
                'date_joined',
                'last_login'
            )
        })
    )
    list_display = (
        'username',
        'first_name',
        'last_name',
        'email',
        'telegram_username'
    )
